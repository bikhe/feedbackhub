import io
from datetime import timedelta

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from users.models import User
from feedback.models import Feedback, Tag
from tasks.models import Task


class ExcelExporter:
    """Экспорт данных в Excel."""

    def __init__(self):
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.positive_fill = PatternFill(start_color='F6FFED', end_color='F6FFED', fill_type='solid')
        self.negative_fill = PatternFill(start_color='FFF2F0', end_color='FFF2F0', fill_type='solid')

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

    def _style_header(self, ws, row=1):
        """Применить стили к заголовкам."""
        for cell in ws[row]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

    def _auto_width(self, ws):
        """Автоширина колонок."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_response(self, wb, filename):
        """Создать HTTP response с файлом."""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def export_feedbacks(self, period_days=30):
        """Экспорт отзывов."""
        now = timezone.now()
        period_start = now - timedelta(days=period_days)

        feedbacks = Feedback.objects.filter(
            created_at__gte=period_start,
        ).select_related(
            'author', 'recipient', 'task'
        ).prefetch_related('tags').order_by('-created_at')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Отзывы'

        # Заголовки
        headers = [
            'Дата',
            'Автор',
            'Email автора',
            'Получатель',
            'Email получателя',
            'Задача',
            'Тональность',
            'Теги',
            'Комментарий',
            'Анонимный',
        ]
        ws.append(headers)
        self._style_header(ws)

        # Данные
        for fb in feedbacks:
            tags_list = ', '.join([t.name for t in fb.tags.all()])
            sentiment = fb.sentiment or 'Н/Д'

            row = [
                fb.created_at.strftime('%Y-%m-%d %H:%M'),
                fb.author.full_name,
                fb.author.email,
                fb.recipient.full_name,
                fb.recipient.email,
                f'[{fb.task.code}] {fb.task.title}',
                'Позитивный' if sentiment == 'positive' else 'Негативный' if sentiment == 'negative' else sentiment,
                tags_list,
                fb.comment,
                'Да' if fb.is_anonymous else 'Нет',
            ]
            ws.append(row)

            # Подсветка по sentiment
            row_num = ws.max_row
            if sentiment == 'positive':
                for cell in ws[row_num]:
                    cell.fill = self.positive_fill
            elif sentiment == 'negative':
                for cell in ws[row_num]:
                    cell.fill = self.negative_fill

        self._auto_width(ws)

        filename = f'feedbacks_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
        return self._create_response(wb, filename)

    def export_users_analytics(self, period_days=30):
        """Экспорт аналитики по пользователям."""
        now = timezone.now()
        period_start = now - timedelta(days=period_days)

        users = User.objects.filter(is_active=True).order_by('last_name')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Аналитика по сотрудникам'

        # Заголовки
        headers = [
            'ФИО',
            'Email',
            'Отдел',
            'Должность',
            'Роль',
            'Получено отзывов',
            'Позитивных',
            'Негативных',
            'Sentiment Score',
            'Оставлено отзывов',
        ]
        ws.append(headers)
        self._style_header(ws)

        # Данные
        for user in users:
            received = Feedback.objects.filter(
                recipient=user,
                created_at__gte=period_start,
            )
            received_total = received.count()
            received_positive = received.filter(tags__sentiment='positive').distinct().count()
            received_negative = received.filter(tags__sentiment='negative').distinct().count()

            given_total = Feedback.objects.filter(
                author=user,
                created_at__gte=period_start,
            ).count()

            total = received_positive + received_negative
            sentiment_score = round((received_positive - received_negative) / max(total, 1) * 100, 1)

            row = [
                user.full_name,
                user.email,
                user.department or '—',
                user.position or '—',
                user.get_role_display(),
                received_total,
                received_positive,
                received_negative,
                sentiment_score,
                given_total,
            ]
            ws.append(row)

        self._auto_width(ws)

        filename = f'users_analytics_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
        return self._create_response(wb, filename)

    def export_summary(self, period_days=30):
        """Экспорт сводной аналитики."""
        now = timezone.now()
        period_start = now - timedelta(days=period_days)

        wb = Workbook()

        # --- Лист 1: Общая статистика ---
        ws1 = wb.active
        ws1.title = 'Общая статистика'

        stats = [
            ['Метрика', 'Значение'],
            ['Период (дней)', period_days],
            ['Всего пользователей', User.objects.filter(is_active=True).count()],
            ['Всего отзывов', Feedback.objects.count()],
            ['Отзывов за период', Feedback.objects.filter(created_at__gte=period_start).count()],
            ['Позитивных за период', Feedback.objects.filter(created_at__gte=period_start, tags__sentiment='positive').distinct().count()],
            ['Негативных за период', Feedback.objects.filter(created_at__gte=period_start, tags__sentiment='negative').distinct().count()],
            ['Всего задач', Task.objects.count()],
            ['Активных задач', Task.objects.filter(status='active').count()],
        ]

        for row in stats:
            ws1.append(row)

        self._style_header(ws1)
        self._auto_width(ws1)

        # --- Лист 2: Топ тегов ---
        ws2 = wb.create_sheet('Топ тегов')

        ws2.append(['Тег', 'Тональность', 'Использований за период', 'Всего использований'])
        self._style_header(ws2)

        tags = Tag.objects.filter(is_active=True).annotate(
            total_usage=__import__('django').db.models.Count('feedbacks'),
            period_usage=__import__('django').db.models.Count(
                'feedbacks',
                filter=__import__('django').db.models.Q(feedbacks__created_at__gte=period_start)
            ),
        ).order_by('-period_usage')

        for tag in tags:
            sentiment = 'Позитивный' if tag.sentiment == 'positive' else 'Негативный'
            ws2.append([
                f'{tag.icon} {tag.name}',
                sentiment,
                tag.period_usage,
                tag.total_usage,
            ])

        self._auto_width(ws2)

        filename = f'summary_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
        return self._create_response(wb, filename)

    def export_full_report(self, period_days=30):
        """Полный отчёт (все данные в одном файле)."""
        now = timezone.now()
        period_start = now - timedelta(days=period_days)

        wb = Workbook()

        # --- Лист 1: Отзывы ---
        ws1 = wb.active
        ws1.title = 'Отзывы'

        headers = ['Дата', 'Автор', 'Получатель', 'Задача', 'Тональность', 'Теги', 'Комментарий']
        ws1.append(headers)
        self._style_header(ws1)

        feedbacks = Feedback.objects.filter(
            created_at__gte=period_start,
        ).select_related(
            'author', 'recipient', 'task'
        ).prefetch_related('tags').order_by('-created_at')

        for fb in feedbacks:
            tags_list = ', '.join([t.name for t in fb.tags.all()])
            sentiment = 'Позитивный' if fb.sentiment == 'positive' else 'Негативный' if fb.sentiment == 'negative' else 'Н/Д'
            ws1.append([
                fb.created_at.strftime('%Y-%m-%d %H:%M'),
                fb.author.full_name,
                fb.recipient.full_name,
                fb.task.code,
                sentiment,
                tags_list,
                fb.comment[:100] + '...' if len(fb.comment) > 100 else fb.comment,
            ])

        self._auto_width(ws1)

        # --- Лист 2: Сотрудники ---
        ws2 = wb.create_sheet('Сотрудники')

        headers = ['ФИО', 'Email', 'Отдел', 'Получено', 'Позитивных', 'Негативных', 'Score', 'Оставлено']
        ws2.append(headers)
        self._style_header(ws2)

        users = User.objects.filter(is_active=True).order_by('last_name')

        for user in users:
            received = Feedback.objects.filter(recipient=user, created_at__gte=period_start)
            received_total = received.count()
            received_positive = received.filter(tags__sentiment='positive').distinct().count()
            received_negative = received.filter(tags__sentiment='negative').distinct().count()
            given_total = Feedback.objects.filter(author=user, created_at__gte=period_start).count()

            total = received_positive + received_negative
            score = round((received_positive - received_negative) / max(total, 1) * 100, 1)

            ws2.append([
                user.full_name,
                user.email,
                user.department or '—',
                received_total,
                received_positive,
                received_negative,
                score,
                given_total,
            ])

        self._auto_width(ws2)

        # --- Лист 3: Теги ---
        ws3 = wb.create_sheet('Теги')

        ws3.append(['Тег', 'Тональность', 'Использований'])
        self._style_header(ws3)

        from django.db.models import Count, Q

        tags = Tag.objects.filter(is_active=True).annotate(
            usage=Count('feedbacks', filter=Q(feedbacks__created_at__gte=period_start))
        ).order_by('-usage')

        for tag in tags:
            ws3.append([
                f'{tag.icon} {tag.name}',
                'Позитивный' if tag.sentiment == 'positive' else 'Негативный',
                tag.usage,
            ])

        self._auto_width(ws3)

        filename = f'full_report_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
        return self._create_response(wb, filename)